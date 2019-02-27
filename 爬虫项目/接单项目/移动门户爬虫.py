# -*- coding:utf-8 -*-
"""
author = zhangql
"""
from openpyxl import Workbook, load_workbook
import requests
import re
from multiprocessing.dummy import Pool


headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/49.0.2623.221 Safari/537.36 SE 2.X MetaSr 1.0"
}
book = Workbook()
sheet = book.active
sheet.append(['开放ID(OPER_ID)', '组织标识(ORG_ID)', '电话(PHONE)', '电话编号(PHONE_NO)', '客户画像', '推荐套餐', '流量', '消费', '通话'])


def get_data(OPER_ID, ORG_ID, PHONE, PHONE_NO):
    url = 'https://smartops.zj.chinamobile.com/MCP/v1/recommendInfo'
    data = {
        'PHONE_NO': PHONE_NO,
        'CHANNEL_ID': '902',
        'CHANNELADIVID': '1',
        'OPER_ID': OPER_ID,
        'ORG_ID': ORG_ID
    }
    try:
        response = requests.post(url, data=data, headers=headers, timeout=25)
        res_json = response.json()
    except Exception as ex:
        return None
    json_list = res_json['customerFigures']
    phone_text = ''
    for i in json_list:
        phone_text += i + '_'

    try:
        phone_traffic = re.findall('\[(.*?)\]', json_list[1], re.S | re.I)[0]  # 上网流量
    except:
        phone_traffic = 0
    try:
        phone_Consumption = re.findall('\[(.*?)\]', json_list[2], re.S | re.I)[0]  # 通话消费
    except:
        phone_Consumption = 0
    try:
        phone_call = re.findall('\[(.*?)\]', json_list[3], re.S | re.I)[0]  # 通话时长
    except:
        phone_call = 0

    if res_json['recommendInfos'] == []:
        P_PRODUCT_NAME = '暂无'
    else:
        P_PRODUCT_NAME = res_json['recommendInfos'][-1]['P_PRODUCT_NAME']
    phone_text = phone_text.rstrip('_')
    print(PHONE, phone_text, P_PRODUCT_NAME)
    sheet.append(
        [OPER_ID, ORG_ID, PHONE, PHONE_NO, phone_text, P_PRODUCT_NAME, phone_traffic, phone_Consumption, phone_call])
    return [OPER_ID, ORG_ID, PHONE, PHONE_NO, phone_text, P_PRODUCT_NAME, phone_traffic, phone_Consumption, phone_call]


def read_excel(file_path):
    file_name = str(file_path).split('xlsx')[0].replace('.', '') + '_new.xlsx'
    workbook = load_workbook(file_path)
    sheets = workbook.get_sheet_names()  # 从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    rows = booksheet.rows
    pool = Pool(processes=100)
    for row in rows:
        line = [col.value for col in row]
        OPER_ID = line[0]
        ORG_ID = line[1]
        PHONE = line[-2]
        PHONE_NO = str(line[-1]).replace(' ', '')
        print(row)
        # pool.apply_async(get_data, (OPER_ID, ORG_ID, PHONE, PHONE_NO))
    pool.close()
    pool.join()
    # book.save(file_name)


if __name__ == '__main__':
    import datetime
    print(datetime.datetime.now())
    # read_excel('/Users/zhangql/python/爬虫项目/接单项目/1.xlsx')
    get_data('20178222', '40122056', '123', '-119,-115,50,22,-23,-49,-3,-43,125,-115,27,31,30,70,-37,74,94,32,-51,-116,115,100,-52,-122,-2,-61,-33,2,112,-36,-103,35,-45,4,65,57,-83,-62,-79,83,47,-44,27,121,-31,-28,-78,47,113,-112,-106,2,89,99,-77,51,-17,117,-104,59,40,102,30,44,111,69,19,12,-18,-77,68,-68,-34,-106,-120,-70,-65,-5,-93,-35,0,97,94,73,101,-84,-105,21,8,87,-6,-113,-100,-6,27,-33,112,46,9,13,-86,53,-46,-69,-66,-30,55,-121,-19,72,-22,78,55,98,-1,33,45,-41,-119,-80,-128,63,-40,45,30,-106,-22,0')
    print(datetime.datetime.now())